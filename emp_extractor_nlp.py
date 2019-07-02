import regex as re
from datetime import datetime
from collections import OrderedDict
import time
import spacy
import openpyxl
from shutil import copyfile


model_directory = r'C:\Users\jhaka\Desktop\intern work\ml_exp_extractor\ner emp extractor\NewNer'
nnlp = spacy.load(model_directory)



CompanyNamePattern = re.compile(r'((company|organi[z|s]ation|employer|Bank)([^a-zA-Z0-9_\']{0,3}name)?\s*[^\w ,.\'\)]{1,3}\s*)',re.IGNORECASE)
DesignationPattern = re.compile(r'\b(analyst|QA|head|professor|lecturer|Product\sOwner(?:s)?|recruiter|Specialist|engineer|Engg(.)?|incharge|operator|executive|manager|consult(?:ant)?|associate|officer|promoter|developer|lead|representative|scientist|senior|se|sr|sale(?:s)?|intern|traine|programmer|(co)?(.)?founder|c.?e.?o|principal(\w)?)\b',re.IGNORECASE)
CompanyPattern = OrderedDict()
CompanyPattern['1'] = re.compile(r"(([S|s]olution\w*|Global|Bank|India|[t|T]echnolog\w*|service\w*|solution\w*)\s*)?(pvt(.)?|private)?\s*\b(ltd(.)?|limited)\b|\b(infotech|llp|inc(\.)?|enterprises|consultancy|agency|(P)\s(Ltd.|ltd|limited))\b", re.IGNORECASE)
CompanyPattern['2'] = re.compile(r""+CompanyNamePattern.pattern+".*[^\r\n$]", re.IGNORECASE)
CompanyPattern['3'] = re.compile(r'\b((employe\w*|work(ing|ed)|deputed|(work|professional|technical)\s*experience|currently associated).*\b(with|as|as a|in|at|for)\b)\b',re.IGNORECASE)
CompanyPattern['4'] = re.compile(r""+DesignationPattern.pattern+".*\s+(at|in)\s+.*[^\r\n$]", re.IGNORECASE)
CompanyPattern['5'] = re.compile(r"(?<=\$\#\$\#\$)[^\#]*(?=\$\#\$\#\$)", re.IGNORECASE)

LatestPattern = re.compile(r'\b(till|current\w*|present\w*|working)\b',re.IGNORECASE)
Delimiters = {"and","role","date","since","behalf","of", "project","work", "client", "not", "till", "present", "employer", "details", "through", "name", "within", "for", "organization", "organisation", "position", "since", "is", "as", "to", "contract", "from", "by", "on", "at", "with", "in", "place", "months"}
DelimitersPattern = re.compile(r'[\d\t\r\n\(\)\[\]\/\:\|\,]|[\s]{4}| ' + (' | ').join(map(re.escape, Delimiters)) + ' ', re.IGNORECASE)
# BlockPattern = re.compile('(?P<pre_lines>(.*[\n]){0,3})(?P<match_lines>(?:(.*(' + ('|').join([ '(%s)' % v.pattern for v in CompanyPattern.values() ]) + ')(.*[\n]){0,3}.*[^\r\n$])))', re.IGNORECASE)
BlockPattern = re.compile('(.*(' + ('|').join([ '(%s)' % v.pattern for v in CompanyPattern.values() ]) + ')(.*[\n]){0,4}.*[^\r\n$])', re.IGNORECASE)
DatePattern = re.compile(r'(\b\d{1,2}\s*[-/:\.]\s*\d{1,2}\s*[-/:\.]\s*\d{2}(?:\d{2})?\b|\b\d{1,2}(?:\')*\s*(?:st|nd|rd|th)*\s*[,-/:\.of]*\s*(?:Jan|january|Feb|February|mar|march|april|Apr|may|june|jun|July|Jul|august|aug|september|sept|sep|oct|october|nov|november|December|dec)+\s*[,-/:\.]*\s*\d{2}(?:\d{2})?\b|\b(?:Jan|january|Feb|February|mar|march|april|Apr|may|june|jun|July|Jul|august|aug|september|sept|sep|oct|october|nov|november|December|dec)+\s*[,-/:\.]*\s*\d{1,2}(?:\')*\s*(?:st|nd|rd|th)*\s*[,-/:\.of]\s*\d{2}(?:\d{2})?\b)',re.IGNORECASE)
MonthYearPattern = re.compile(r'(\b(?:Jan|january|Feb|February|mar|march|april|Apr|may|june|jun|July|Jul|august|aug|september|sept|sep|oct|october|nov|november|December|dec)+\s*[,-/:\.\']*\s*\d{2}(?:\d{2})?|\s*[,-/:\.]\d{1,2}\s*[,-/:\.]\s*\d{2}(?:\d{2})?|\s*[,-/:\.\s]\d{1,2}\s*[,-/:\.]\s*\d{4}\b)',re.IGNORECASE)
MonthYearPattern2 = re.compile(r'(\b(?:Jan|january|Feb|February|mar|march|april|Apr|may|june|jun|July|Jul|august|aug|september|sept|sep|oct|october|nov|november|December|dec)+\s*[,-/:\.\']*\s*\d{2}(?:\d{2})?(?:\s*[,-/:\.]\s*\d{2}(?:\d{2})?\b))',re.IGNORECASE)

OnlyYearPattern = re.compile(r'\b([1|2]{1}[9|0]{1}[0-9]{2})\b',re.IGNORECASE)
CompBlockPattern = re.compile(r'\b(tech|service|solution|System|Corporate|Market|bank)\b',re.IGNORECASE)
month_reg = re.compile(
            r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)(\w*[a-zA-z]+)?', re.IGNORECASE)
month_dict = {
            "jan": "01",
            "feb": "02",
            "mar": "03",
            "apr": "04",
            "may": "05",
            "jun": "06",
            "jul": "07",
            "aug": "08",
            "sep": "09",
            "oct": "10",
            "nov": "11",
            "dec": "12"
        }

EmployerLmt = 50
DesignationLmt = 50

class WorkProfile(object):

    def __init__(self):
        self.company = ''
        self.location = ''
        self.start = 0
        self.end = 0
        self.designation = ''
        self.latest = False
        self.ccount = 0
        self.lcount = 0
        self.dcount = 0
        self.dacount = 0
        self.case = ''



class EmployerExtractor:

    def __init__(self, z):
        self.exp = []
        '''z = re.sub(r'[^\x00-\x7F]+',' ', z)
        z = re.sub(r'[^\S\n]+', ' ', re.sub(r'\s*[\n\t\r]\s*', '\n', z))
        z=re.sub('’','\'',z)
        z=z.strip('\n')'''
    
        self.text = z
        self.latest_found = False
        self.clients = []
        self.count_dict = {} #{emp:{count, index}}x

    def Extract(self):
        try:
            self.LevelOneMatch()
        except:
            pass
        # return [ e.__dict__ for e in self.exp ]
        return self.CleanEmployer()

    def LevelOneMatch(self):   
            #f=open("resumeresult.txt","a+")
            doc_to_test=nnlp(self.text)
            d = []
            d_dict = {}
            cnt = 0;
            for ent in doc_to_test.ents:
                d_dict[ent.start] = ent
                d.append(ent.start)
                cnt += 1
            #print(d)
            #print(d_dict)
            if len(d)>1:
                d.sort()
                
                diff = [y - x for x, y in zip(*[iter(d)] * 2)]
                avg = sum(diff) / len(diff)
                #print(avg)
                m = [[d[0]]]
                
                for x in d[1:]:
                    if x - m[-1][0] < avg:
                        m[-1].append(x)
                    else:
                        m.append([x])
                
                
                #print(m)
                groups = []
                for i in m:
                    li=[]
                    for x in i:
                        li.append(d_dict[x])
                    groups.append(li)
                #print(groups)
            else:
                groups = [[ent for ent in doc_to_test.ents]]
            for group in groups:
                this_exp = WorkProfile()
                for ent in group:
                    #print(ent.label_,ent,'-',ent.start)
                    if ent.label_ == 'company':
                        this_exp.ccount += 1
                        #print(this_exp.ccount,this_exp.dcount,this_exp.dacount)
                        if this_exp.ccount>1:
                            self.exp.append(this_exp)
                            this_exp = WorkProfile()
                            this_exp.ccount = 1
                        this_exp.company = str(ent)
                    
                    if ent.label_ == 'location':
                        this_exp.lcount += 1
                        #print(this_exp.ccount,this_exp.dcount,this_exp.dacount)
                        if this_exp.lcount>1:
                            self.exp.append(this_exp)
                            this_exp = WorkProfile()
                            this_exp.lcount = 1
                        this_exp.location = str(ent)
                        
                    if ent.label_ == 'designation':
                        this_exp.dcount += 1
                        designation = self.GetDesignation(str(ent))
                                
                        #print(this_exp.ccount,this_exp.dcount,this_exp.dacount)
                        if this_exp.dcount>1:
                                    self.exp.append(this_exp)
                                    this_exp = WorkProfile()
                                    this_exp.dcount = 1
                        this_exp.designation = designation
                        
                    if ent.label_ == 'duration':
                        this_exp.dacount += 1
                        start, end, is_latest = self.GetDuration(str(ent))
                        if start:
                                
                                #print(this_exp.ccount,this_exp.dcount,this_exp.dacount)
                                if this_exp.dacount>1:
                                    self.exp.append(this_exp)
                                    this_exp = WorkProfile()
                                    this_exp.dacount = 1
                                this_exp.start = start
                                if end:
                                    this_exp.end = end
                                elif is_latest:
                                    this_exp.latest = is_latest
                    
                self.exp.append(this_exp)
            
        


    def GetDuration(self, line):
        start, end, is_latest = 0, 0, 0
        # month_year = re.findall(DatePattern, line)
        # if not month_year:
        #print(line)
        count = 0
        month_year = re.findall(MonthYearPattern2, line)
        if not month_year:
            month_year = re.findall(MonthYearPattern, line)
        else:
            count = 1
        #print(month_year)
        if len(month_year) >= 2:
                start = month_year[0]
                end = month_year[1]
        elif len(month_year) == 1:
            start = month_year[0]
            is_latest_list = re.findall(LatestPattern, line)
            if len(is_latest_list) and (not self.latest_found):
                is_latest = True
                self.latest_found = True
        if not start:
            years = re.findall(OnlyYearPattern, line)
            if len(years) >= 2:
                start = '01/%s' % years[0]
                end = '01/%s' % years[1]
            elif len(years) == 1:
                start = '01/%s' % years[0]
                is_latest_list = re.findall(LatestPattern, line)
                if len(is_latest_list) and (not self.latest_found):
                    is_latest = True
                    self.latest_found = True
        if start:
            start = self.CleanDate(start,count)
        if end:
            end = self.CleanDate(end,count)
        return start, end, is_latest

    def GetDesignation(self, line):
        #print(line,' 999')
        return_val = ''
        if len(line)>200:
            return ''
        rr=re.compile(r'(?<=(\bdesignation|role|position\b\W+)).*',re.IGNORECASE)
        match = re.search(rr, line)
                                        
        if match:
            designation = re.sub(r'^(\W*(an|a)?\W+)', '', match.group())
            designation = self.CleanStringName(designation)
            if (len(designation) > 2 and len(designation) < DesignationLmt and len(designation.split())<4):
                return_val = designation
        if not return_val and re.search('work',line,re.IGNORECASE):
            match = re.search(r'(?<=(\bas\b)).*', line)
            if match:
                designation = re.sub(r'^(\W*(an|a)?\W+)', '', match.group())
                splits = re.split(DelimitersPattern, designation)
                designation = self.CleanStringName(splits[0])
                #print(designation,'::::::::::::::::::::')
                if (len(designation) > 2 and len(designation) < DesignationLmt and len(designation.split()) < 4):
                    return_val = designation
        match = re.search(DesignationPattern, line)
        #print(match,'555555555555555555555555555555555555')
        if not return_val and match:
            
                    designation = self.CleanStringName(line)
                    return_val = designation
        return return_val

    def CleanStringName(self, name, is_company=False):
        #print(name,'company')
        #print(is_company,'000000000000')
        
        if is_company:
            name = name.strip()
            name = name[::-1]
            match = re.findall(r'[^0-9a-zA-Z\.\&\,\- ]', name)
            
            if len(match) and name.index(match[0]) > 2:
                name = name[0:name.index(match[0])]
                #print(name,'858585858558585')
            elif len(match) > 1 and name.index(match[1]) > 2:
                name = name[0:name.index(match[1])]
            name = re.sub('\s+', ' ', name).strip('., ')
            #print(name,'123456789')
            if is_company:
                name = name[::-1]
        else:
            name = re.sub(r'\*|\.|\?|\||\/|\:|\(|\)','-',name)
            name = re.sub(r'\+','\+',name)
            
            splitt=''
            match = re.findall(r'[^\w ]', name)
            if match:
                if len(match)>1:
                    for i in match:
                        splitt+=i+'|'
                    splitt=splitt.strip('|')
                else:
                    splitt=match[0]
                desig_splits=re.split(splitt,name)
                for desig in desig_splits:
                    if re.search(DesignationPattern,desig):
                        name=desig
                
                
        return name

    def CleanDate(self, _date,count):
        _now = datetime.now()
        _date = _date.lower().strip()
        try:
            r = '' + month_dict[re.search(month_reg, _date).group()[0:3]] + '/'
            _date = re.sub(month_reg, r, _date)
        except:
            pass
        _date = re.sub(r'([^0-9a-z])+', '/', _date).strip('/')
        #print(_date)
        if count == 1:
            month = _date.split("/")[0]
            year = _date.split("/")[2]
        else:
            month = _date.split("/")[0]
            year = _date.split("/")[1]
        #print(month,year)
        if len(year) == 2:
            if int(year) < 50:
                year = "%s%s" % ('20', year)
            else:
                year = "%s%s" % ('19', year)
        if len(month) == 1:
            if int(month) < 10:
                month = "%s%s" % ('0', month)
        _date = "%s/%s" % (month, year)
        if int(year) > _now.year:
            _date = None
        return _date

    def CleanEmployer(self):
        data = []
        count=0
        for e in self.exp:
            if e.company.title() or e.designation.title():
                this_data = OrderedDict()
                this_data['Company'] = e.company.title()
                this_data['Location'] = e.location.title()
                this_data['Designation'] = e.designation.title()
                this_data['From'] = e.start
                this_data['FromMonth'] = e.start.split('/')[0] if e.start else None
                this_data['To'] = e.end
                this_data['ToMonth'] = e.end.split('/')[0] if e.end else None
                this_data['IsLatest'] = e.latest if self.latest_found else (not count)
                data.append(this_data)
                count+=1
        return data


if __name__ == "__main__":
    
    book=openpyxl.load_workbook('check.xlsx')
    sheet=book.get_sheet_by_name('Sheet1')
    ro=1  
    toRem = []
    designation = []
    for cnt in range(1,26000):
        filepath = r'C:\Users\jhaka\Desktop\intern work\all resumes\text_resume\resumee '+str(cnt)+'.txt'  #path to the text resume 
        
        with open(filepath,'r') as f:
            z=f.read()
            
        start = time.clock()
        x = EmployerExtractor(z)
        data = x.Extract()
        print('resume',cnt)
        #print(time.clock() - start)
        #print(data)
        if not data:
            copyfile(filepath, r'C:\Users\jhaka\Downloads\resumes-for-AI\text_resume\remove\resume'+str(cnt)+'.txt')
        for v in data:
            try:
                com=v['Company']
                loc = v['Location']
                des=v['Designation']
                fr=v['From']
                t=v['To']

            except KeyError:
                pass
            li=[]
            #li.append(com)
            #li.append(loc)
            li.append(des)
            if des. not in designation and des:
                designation.append(des)
            #li.append(fr)
            #li.append(t)
            li.append('resume'+str(cnt))

            for n in range(1,len(li)+1):
               sheet.cell(row=ro,column=n).value=li[n-1]
            ro+=1
            #book.save('check.xlsx')
        ro+=1

designation.sort(key = lambda s: len(s.strip()),reverse = True) 
book=openpyxl.load_workbook('ner-designa.xlsx')
sheet=book.get_sheet_by_name('Sheet1')
ro=1

for x in designation:
   print(ro)
   sheet.cell(row=ro,column=1).value=x.strip()
   ro+=1
   book.save('ner-designa.xlsx')

# =============================================================================
# x = 25,50,60
# 
# lis = list()
# 
# print(lis.append(4))
# 
# def myFun(*args):  
#     print(args)
# 
# myFun('a',1,x)
# 
# def myFun(**kwargs):  
#     for key, value in kwargs.items():
#         print ("%s == %s" %(key, value)) 
#   
# # Driver code 
# myFun(first ='Geeks', mid ='for', last='Geeks')   
# 
# try:
#     if '1' != 1:
#         raise "someError"
#     else:
#         print("someError has not occured")
# except "someError":
#     print ("someError has occured")
# 
# 
# 
# li=['aman','aman kr']
# for x in enumerate(li):
#     print(x)
# my_tuple = (3, 8, 8, 8) 
# my_tuple1 = (5, 2, 3, 4)
# print (my_tuple > my_tuple1)
# 
# a = 'aa'*10
# b = 'a'*20
# print(id(a)==id(b))
# 
# a = "aman"
# b = "aman"
# 
# i = 1
# while True: 
#     if i % 0O7 == 0: 
#         break
#     print(i) 
#     i += 1
#     
# class a:
#     x=10
# 
# class b:
#     x=20
# 
# class c(a,b):
#     def __init__(self):
#         print(self.x)
#         
# print(c())
#     
#     
# import timeit
# 
# example = [i for i in range(100)]
# 
# for i in range(len(example)):
#     x = i
#     y = example[i]
# 
# 
# timeit.timeit('''example = [i for i in range(100000)]
# 
# for i in range(len(example)):
#     x = i
#     y = example[i]''', number=500)
#     
# 
# timeit.timeit('''example = [i for i in range(100000)]
# 
# for i,j in enumerate(example):
#     x = i
#     y = j''', number=500)
#     
# 
# list1 = ['a', 'b', 'c', 'd']
# list1[0:2] = 'z'
# print(list1)
# sum = lambda x,y,z=0: (x+y+z)
# sum(2,2)
# 
# 
# li = range(10)
# final_list = list(filter(lambda x: (x%2 != 0) , li)) 
# print(final_list)
# 
# final_list = [i for i in range(10) if i%2!=0]
# 
# timeit.timeit('''final_list =[i for i in range(10) if i%2!=0]''', number=500000)
# 
# y=[1,2,3]
# print(id(y))
# def test(y):
#     print(id(y))
#     y += [1]    #y.__iadd__([1])        x += 1 is different than x = x + 1 bez += is inplace operation n others are just reasignment
#     print(id(y))
# test(y)
# print(id(y))
# print(y)
# 
# timeit.timeit('''[x for x in range(100000) if x%2==1]''', number=500)
# 
# timeit.timeit('''list(filter(lambda x: x % 2 == 1, range(100000)))''', number=500)
# print(len(list(filter(lambda x: x % 2 == 1, range(100000)))))
# 
# 
# def fun(x):
#     print(id(x),'in')
#     x=[5]
#     print(id(x))
#     print(x)
# x=300
# print(id(x),'out')
# fun(x)
# print(id(x))
# print(x)
# 
# x = [12,23]
# y = x.copy
# print(id(x),id(y))
# =============================================================================
